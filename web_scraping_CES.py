"""
Web Scraping in CES site to generate the exhibitor list
"""

import openpyxl as excel
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time

# the option can work on any environment without opening web browser
op = Options()
op.add_argument("--disable-gpu")
op.add_argument("--disable-extensions")
op.add_argument("--proxy-server='direct://'")
op.add_argument("--proxy-bypass-list=*")
op.add_argument("--start-maximized")
op.add_argument("--headless")

# Parameters for excel

# To change the file name
saveName = 'ces2020.xlsx'
# To change the sheet name of the above file name
sheetTitle = 'CES2020'

# Parameters for urls

# To change URL access
url1 = 'https://www.ces.tech/Show-Floor/Exhibitor-Directory.aspx?searchTerm=&sortBy=alpha&filter='
url2 = '&pageNo=1&pageSize=30'

# to get by all categories or each alphabet
# features = '#|ABCDEFGHIJKLMNOPQRSTUVWXYZ'
features = '#'


def scroll(driver, timeout):
    """
    Auto scrolling to the bottom
    """
    scroll_pause_time = timeout

    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        # Scroll down to bottom
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load page
        time.sleep(scroll_pause_time)

        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            # If heights are the same it will exit the function
            break
        last_height = new_height


def getData(url):
    """
    start the session to the url provided, to get and store
    the data in excel sheet from url
    """
    # wait for the driver ready. It require to avoide the connection refused
    time.sleep(10)
    # chromedriver is required in /driver folder
    driver = webdriver.Chrome(options=op)

    # Start session
    driver.get(url)

    # Call Auto Scroll
    scroll(driver, 5)

    # Scraping
    names = driver.find_elements_by_css_selector('.company-name')
    companyName = [name.text for name in names]
    lenName = len(companyName)

    booths = driver.find_elements_by_css_selector('.bd-highlight footer')
    companyBooth = [booth.text for booth in booths]
    lenBooth = len(companyBooth)

    urls = driver.find_elements_by_css_selector('.company-name a')
    companyUrl = [url.get_attribute('href') for url in urls]
    lenUrl = len(companyUrl)

    # close the browser
    driver.quit()

    # Data to Excel
    sheet, max_row, wb = checkExcel()
    inputRow = max_row + 1
    if lenName == lenBooth and lenName == lenUrl:
        for i in range(lenName):
            sheet['A' + str(i + inputRow)] = companyName[i]
            sheet['B' + str(i + inputRow)] = companyUrl[i]
            sheet['C' + str(i + inputRow)] = companyBooth[i].replace(
                'BOOTH:', '').strip()
        # save the data in the excel
        wb.save(saveName)
    else:
        print('Length is different between name, booth and url')


def checkExcel():
    """
    To create if the file is not existed, otherwise to open the existing file
    """
    if not os.path.exists(saveName):
        max_row = 1
        wb = excel.Workbook()
        sheet = wb.active
        sheet.title = sheetTitle
        sheet['A1'] = 'Company'
        sheet['B1'] = 'CES URL'
        sheet['C1'] = 'Booth#'
    else:
        wb = excel.load_workbook(saveName)
        sheet = wb.active
        max_row = wb[sheetTitle].max_row
    return (sheet, max_row, wb)


for i in range(len(features)):
    print('start: ' + features[i])
    url = url1 + features[i] + url2

    # get the data
    getData(url)
